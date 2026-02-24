"""Tests for excel_export module."""

from io import BytesIO

import openpyxl
import pandas as pd
import pytest

from excel_export import generate_excel, num_to_letter


# --- num_to_letter ---


class TestNumToLetter:
    def test_zero_returns_a(self):
        assert num_to_letter(0) == "A"

    def test_sequential(self):
        assert num_to_letter(1) == "B"
        assert num_to_letter(2) == "C"
        assert num_to_letter(25) == "Z"


# --- generate_excel ---


@pytest.fixture
def excel_inputs():
    """Minimal DataFrames for generate_excel."""
    left = pd.DataFrame(
        {"Werknemer 1": ["Alice"], "Werknemer 2": ["Bob"]},
        index=["Ompakken"],
    )
    right = pd.DataFrame(
        {"Werknemer 1": ["Charlie"]},
        index=["Krimpen"],
    )
    full = pd.DataFrame(
        {0: ["Alice", "Charlie"], 1: ["Bob", ""]},
        index=["Ompakken", "Krimpen"],
    )
    full.columns = full.columns.astype(str)
    afwezig = pd.DataFrame({0: ["Wim"]})
    return left, right, full, afwezig


class TestGenerateExcel:
    def test_returns_bytes(self, excel_inputs):
        left, right, full, afwezig = excel_inputs
        result = generate_excel(left, right, full, afwezig, "maandag",
                                logo_url=None, friday_image_url=None)
        assert isinstance(result, bytes)

    def test_output_is_valid_xlsx(self, excel_inputs):
        left, right, full, afwezig = excel_inputs
        result = generate_excel(left, right, full, afwezig, "maandag",
                                logo_url=None, friday_image_url=None)
        wb = openpyxl.load_workbook(BytesIO(result))
        assert wb is not None
        wb.close()

    def test_planning_sheet_exists(self, excel_inputs):
        left, right, full, afwezig = excel_inputs
        result = generate_excel(left, right, full, afwezig, "maandag",
                                logo_url=None, friday_image_url=None)
        wb = openpyxl.load_workbook(BytesIO(result))
        assert "Planning" in wb.sheetnames
        wb.close()

    def test_task_names_in_output(self, excel_inputs):
        left, right, full, afwezig = excel_inputs
        result = generate_excel(left, right, full, afwezig, "maandag",
                                logo_url=None, friday_image_url=None)
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb["Planning"]
        all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "Ompakken" in all_values
        wb.close()

    def test_friday_variant(self, excel_inputs):
        left, right, full, afwezig = excel_inputs
        result = generate_excel(left, right, full, afwezig, "vrijdag",
                                logo_url=None, friday_image_url=None)
        assert isinstance(result, bytes)

    def test_absent_workers_in_output(self, excel_inputs):
        left, right, full, afwezig = excel_inputs
        result = generate_excel(left, right, full, afwezig, "maandag",
                                logo_url=None, friday_image_url=None)
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb["Planning"]
        all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "Wim" in all_values
        wb.close()

    def test_empty_afwezig(self, excel_inputs):
        left, right, full, _ = excel_inputs
        afwezig = pd.DataFrame({0: []})
        result = generate_excel(left, right, full, afwezig, "maandag",
                                logo_url=None, friday_image_url=None)
        assert isinstance(result, bytes)
